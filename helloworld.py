from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font
# We need openpyxl library to be able to create, read and write on excel
# .styles is used in this case to manipulate if the Font is bold or normal

class ContactManager:
    def __init__(self, filename):
        self.filename = filename
        try:
            self.workbook = load_workbook(filename)
        except FileNotFoundError:
            self.workbook = Workbook()
        self.sheet = self.workbook.active

        # Define headers if they don't exist
        headers = ["F.Name", "L.Name", "Phone", "Birthday.D", "Birthday.M", "Birthday.Y"]
        if self.sheet.max_row == 0 or any(self.sheet.cell(row=1, column=col+1).value != header
                                          for col, header in enumerate(headers)):
            self.add_headers(headers)

    #If headers already exist
    def add_headers(self, headers):
        for col, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)

    def add_contact(self, first_name, last_name, phone, birthday_month, birthday_day, birthday_year):
        next_row = self.sheet.max_row + 1
        self.sheet.cell(row=next_row, column=1, value=first_name)
        self.sheet.cell(row=next_row, column=2, value=last_name)
        self.sheet.cell(row=next_row, column=3, value=phone)
        self.sheet.cell(row=next_row, column=4, value=birthday_month)
        self.sheet.cell(row=next_row, column=5, value=birthday_day)
        self.sheet.cell(row=next_row, column=6, value=birthday_year)
        self.workbook.save(self.filename)

    def list_contacts(self):
        for row in range(2, self.sheet.max_row + 1):
            first_name = self.sheet.cell(row=row, column=1).value
            last_name = self.sheet.cell(row=row, column=2).value
            phone = self.sheet.cell(row=row, column=3).value
            birthday_month = self.sheet.cell(row=row, column=4).value
            birthday_day = self.sheet.cell(row=row, column=5).value
            birthday_year = self.sheet.cell(row=row, column=6).value
            print(f"First Name: {first_name}, Last Name: {last_name}, Phone: {phone}, Birthday: {birthday_month} {birthday_day} {birthday_year}")

    def edit_contact(self, first_name, new_phone, new_birthday_month, new_birthday_day, new_birthday_year):
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value == first_name:
                self.sheet.cell(row=row, column=3, value=new_phone)
                self.sheet.cell(row=row, column=4, value=new_birthday_month)
                self.sheet.cell(row=row, column=5, value=new_birthday_day)
                self.sheet.cell(row=row, column=6, value=new_birthday_year)
                self.workbook.save(self.filename)
                print(f"Contact '{first_name}' updated successfully.")
                return
        print(f"Contact '{first_name}' not found.")

    def search_contact(self, first_name):
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value == first_name:
                last_name = self.sheet.cell(row=row, column=2).value
                phone = self.sheet.cell(row=row, column=3).value
                birthday = self.sheet.cell(row=row, column=4).value
                print(f"Name: {first_name}, Last Name: {last_name}, Phone: {phone}, Birthday: {birthday}")
                return
        print(f"Contact '{first_name}' not found.")

    def delete_contact(self, first_name):
        for row in range(2, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value == first_name:
                self.sheet.delete_rows(row, amount=1)
                self.workbook.save(self.filename)
                print(f"Contact '{first_name}' deleted successfully.")
                return
        print(f"Contact '{first_name}' not found.")

def validate_empty(ambiguos_input):
    if ambiguos_input == "":
        print("Please, enter valid input")
        return True
    else:
        return False

# start of the code
contact_manager = ContactManager("contacts.xlsx")

while True:
    print("\nContact Manager")
    print("1. Add Contact")
    print("2. Edit Contact")
    print("3. Search Contact")
    print("4. Delete Contact")
    print("5. List Contacts")
    print("6. Exit")
    choice = input("Enter your choice (1-6): ")

    match choice:
        case "1":
            while True:
                first_name = input("Enter first name: ")
                if not(validate_empty(first_name)) and (first_name.isalpha()):
                    break
            # don't pass until != empty
            while True:
                last_name = input("Enter last name: ")
                if not(validate_empty(last_name)) and (last_name.isalpha()):
                    break
            # don't pass until != empty
            while True:
                phone = input("Enter phone(0000000000): ")
                if not(validate_empty(phone)) and (phone.isdigit()):
                    break
            # don't pass until != empty
            while True:
                birthday_month = input("Enter birthday month(MM): ")
                if not(validate_empty(birthday_month)) and (birthday_month.isdigit()):
                    break
            while True:
                birthday_day = input("Enter birthday day(DD): ")
                if not(validate_empty(birthday_day)) and (birthday_day.isdigit()):
                    break
            while True:
                birthday_year = input("Enter birthday year(YYYY): ")
                if not(validate_empty(birthday_year)) and (birthday_year.isdigit()):
                    break
            # don't pass until != empty
            # contact_manager.phone_validation(phone)
            contact_manager.add_contact(first_name, last_name, phone, birthday_month, birthday_day, birthday_year)
        case "2":
            first_name = input("Enter first name of contact to edit: ")
            while True:
                new_phone = input("Enter new phone(0000000000): ")
                if not(validate_empty(new_phone)) and (new_phone.isdigit()):
                    break
            while True:
                new_birthday_month = input("Enter new birthday month(MM): ")
                if not(validate_empty(new_birthday_month)) and (new_birthday_month.isdigit()):
                    break
            while True:
                new_birthday_day = input("Enter new birthday day(DD): ")
                if not(validate_empty(new_birthday_day)) and (new_birthday_day.isdigit()):
                    break
            while True:
                new_birthday_year = input("Enter new birthday year(YYYY): ")
                if not(validate_empty(new_birthday_year)) and (new_birthday_year.isdigit()):
                    break
            contact_manager.edit_contact(first_name, new_phone, new_birthday_month, new_birthday_day, new_birthday_year)
        case "3":
            while True:
                first_name = input("Enter first name to search: ")
                if not(validate_empty(first_name)) and (first_name.isalpha()):
                    break
            contact_manager.search_contact(first_name)
        case "4":
            while True:
                first_name = input("Enter first name to delete: ")
                if not(validate_empty(first_name)) and (first_name.isalpha()):
                    break
            contact_manager.delete_contact(first_name)
        case "5":
            contact_manager.list_contacts()
        case "6":
            break
        case _:
            print("Invalid choice. Please try again.")